from openpyxl import load_workbook
from datetime import date, datetime
import os
import shutil


def format_date(val):
    '''Функция для формата ячеек с датами.'''

    return val.strftime('%d.%m.%Y') if isinstance(val, date) else str(val or '')


data_file = 'я. Бетон (Июль).xlsx'  # Путь к файлу с данными
template_file = 'Шаблон ГИ.xlsx'  # Путь к шаблону для актов
output_folder = 'Акты_ги'  # Название для папки под акты

os.makedirs(output_folder, exist_ok=True)  # Создаем папку для выходных файлов

wb_data = load_workbook(data_file)
ws_data = wb_data['ГИ для АОСР']

wb_template = load_workbook(template_file)
ws_template = wb_template['АОСР ГИ']

# dkbs = 'документ о качестве бетонной смеси заданного состава качества партии'
# name_uzk = 'Протокол оценки прочности бетона монолитных железобетонных конструкций'
name_k1 = 'Протокол определения фактической влажности бетонного основания при устройстве гидроизоляции'
name_k2 = 'Протокол определения адгезии гидроизоляционного покрытия'

for row in ws_data.iter_rows(min_row=3, values_only=True):
    '''Сохраняем данные из файла exel в переменные. '''

    id = row[0]
    act_number = str(row[1])
    work_name = str(row[2])
    start_date = format_date(row[3])
    end_date = format_date(row[4])
    next_work = str(row[5]) if row[5] else ''
    material = str(row[7]) if row[7] else ''
    material_data = str(row[8]) if row[8] else ''
    # lab_uzk = str(row[10]) if row[10] else ''
    lab_k = str(row[11]) if row[11] else ''
    lab_date = format_date(row[12])
    code = str(row[13])
    agreement_date = format_date(row[14])

    new_file = os.path.join(output_folder, f'Акт_№{id}.xlsx')
    shutil.copy(template_file, new_file)

    wb_new = load_workbook(new_file)
    ws_new = wb_new['АОСР ГИ']

    # Проверка даты акта
    act_date = max(
        (row[4].date() if isinstance(row[4], datetime) else row[4]) if isinstance(row[4], date) else date.min,
        (row[12].date() if isinstance(row[12], datetime) else row[12]) if isinstance(row[12], date) else date.min,
        (row[14].date() if isinstance(row[14], datetime) else row[14]) if isinstance(row[14], date) else date.min
    ).strftime('%d.%m.%Y')

    # Проверка ЖАН
    agreement = f'Запись из ЖАН от {agreement_date}' if agreement_date else ''

    # Проверка материалов, реестр или нет
    # if mixture_number == 'Реестр':
    #     material1 = f'Материалы согласно реестру №{act_number} от {end_date}'
    #     material2 = f'Реестр №{act_number} от {end_date}'
    #     material1_1, material2_1 = '', ''
    if '\n' in material:
        material_number_parts = material.split('\n')
        material_date_parts = material_data.split('\n')
        material1 = f'{material_number_parts[0]} - {material_date_parts[0]}'
        material1_1 = f'{material_number_parts[1]} - {material_date_parts[1]}'
        material2 = material_date_parts[0].capitalize()
        material2_1 = material_date_parts[1].capitalize()
    else:
        material1 = f'{material} - {material_data}'
        material2 = material_data.capitalize()
        material1_1, material2_1 = '', ''

    # Проверка лаборатории, УЗК или К
    if not lab_k:
        lab1 = ''
        lab2 = ''
    else:
        lab1 = f'{name_k1} №{lab_k}-ВЛ/2/1.3В-2025 от {start_date}'
        lab2 = f'{name_k2} №{lab_k}-А/2/1.3В-2025 от {lab_date}'

    if not next_work:
        next_work = 'Согласно проекта'

    # Список соответсвия мест в exel и переменных
    replacements = {
        '[№ акта]': act_number,
        '[Наименование работ]': work_name,
        '[Дата начала работы]': start_date,
        '[Дата окончания работы]': end_date,
        '[Шифр]': code,
        '[Согласование]': agreement,
        '[Материалы1]': material1,
        '[Материалы2]': material2,
        '[Материалы1_1]': material1_1,
        '[Материалы2_1]': material2_1,
        '[Лаборатория1]': lab1,
        '[Лаборатория2]': lab2,
        '[Дата акта]': act_date,
        '[Следующая работа]': next_work
    }

    # Заполнение ячеек в новом документе
    for cell in ws_new:
        for c in cell:
            if c.value is not None:
                for placeholder, value in replacements.items():
                    if placeholder in str(c.value):
                        c.value = c.value.replace(placeholder, value)

    # Скрытие строк по известным номерам
    rows_to_hide = []

    if not material1_1:
        rows_to_hide.append(76)
    if not material2_1:
        rows_to_hide.append(97)
    if not lab_k:
        rows_to_hide.append(80)
        rows_to_hide.append(81)
        rows_to_hide.append(98)
        rows_to_hide.append(99)
    if not agreement_date:
        rows_to_hide.append(100)

    for row_num in rows_to_hide:
        ws_new.row_dimensions[row_num].hidden = True

    wb_new.save(new_file)
