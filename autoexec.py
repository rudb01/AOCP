from openpyxl import load_workbook, Workbook
from datetime import date, datetime
import os
import logging
import sys
from copy import copy


def setup_logging():
    '''Настройка логирования для отслеживания процесса выполнения.'''
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('autoexec.log', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )


def format_date(val):
    '''Функция для формата ячеек с датами.'''
    return (val.strftime('%d.%m.%Y') if isinstance(val, date)
            else str(val or ''))


def validate_files(data_file, template_file):
    '''Проверка существования необходимых файлов.'''
    if not os.path.exists(data_file):
        raise FileNotFoundError(f"Файл данных '{data_file}' не найден")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Файл шаблона '{template_file}' не найден")
    logging.info(f"Файлы проверены: {data_file}, {template_file}")


def get_act_date(*dates):
    '''Упрощенная функция для определения даты акта.'''
    valid_dates = []
    for dt in dates:
        if dt:
            if isinstance(dt, datetime):
                valid_dates.append(dt.date())
            elif isinstance(dt, date):
                valid_dates.append(dt)
    if valid_dates:
        return max(valid_dates).strftime('%d.%m.%Y')
    return date.today().strftime('%d.%m.%Y')


def copy_print_settings(source_ws, target_ws):
    '''Копирует настройки печати из исходного листа в целевой.'''
    try:
        # Копируем область печати
        if source_ws.print_area:
            # Заменяем имя листа в области печати
            print_area = source_ws.print_area
            if '!' in print_area:
                # Заменяем имя листа на имя целевого листа
                sheet_part, range_part = print_area.split('!', 1)
                target_ws.print_area = f"'{target_ws.title}'!{range_part}"
            else:
                target_ws.print_area = print_area

        # Копируем заголовки печати
        if source_ws.print_titles:
            target_ws.print_titles = source_ws.print_titles

        # Копируем настройки страницы
        target_ws.page_setup.orientation = source_ws.page_setup.orientation
        target_ws.page_setup.paperSize = source_ws.page_setup.paperSize
        target_ws.page_setup.fitToHeight = source_ws.page_setup.fitToHeight
        target_ws.page_setup.fitToWidth = source_ws.page_setup.fitToWidth
        target_ws.page_setup.scale = source_ws.page_setup.scale

        # Копируем отступы страницы
        target_ws.page_margins.left = source_ws.page_margins.left
        target_ws.page_margins.right = source_ws.page_margins.right
        target_ws.page_margins.top = source_ws.page_margins.top
        target_ws.page_margins.bottom = source_ws.page_margins.bottom
        target_ws.page_margins.header = source_ws.page_margins.header
        target_ws.page_margins.footer = source_ws.page_margins.footer

        # Копируем заголовки строк и столбцов для печати
        if source_ws.print_title_rows:
            target_ws.print_title_rows = source_ws.print_title_rows
        if source_ws.print_title_cols:
            target_ws.print_title_cols = source_ws.print_title_cols

        # Копируем настройки представления (вид отображения)
        if source_ws.sheet_view:
            source_view = source_ws.sheet_view
            target_view = target_ws.sheet_view

            # Копируем основные настройки представления
            if source_view.view:
                target_view.view = source_view.view
            if source_view.zoomScale:
                target_view.zoomScale = source_view.zoomScale
            if source_view.zoomScaleNormal:
                target_view.zoomScaleNormal = source_view.zoomScaleNormal
            if source_view.zoomScalePageLayoutView:
                target_view.zoomScalePageLayoutView = \
                    source_view.zoomScalePageLayoutView
            if source_view.showGridLines is not None:
                target_view.showGridLines = source_view.showGridLines
            if source_view.showRowColHeaders is not None:
                target_view.showRowColHeaders = source_view.showRowColHeaders

        # Копируем настройки центрирования при печати
        if source_ws.print_options.horizontalCentered is not None:
            target_ws.print_options.horizontalCentered = \
                source_ws.print_options.horizontalCentered
        if source_ws.print_options.verticalCentered is not None:
            target_ws.print_options.verticalCentered = \
                source_ws.print_options.verticalCentered

        msg = f"Настройки печати, представления и центрирования " \
              f"скопированы для листа '{target_ws.title}'"
        logging.info(msg)

    except Exception as e:
        logging.warning(f"Ошибка при копировании настроек печати: {e}")
        # Продолжаем выполнение, даже если настройки печати не скопировать


def copy_worksheet(source_ws, target_wb, sheet_name):
    '''Копирует лист с сохранением форматирования и настроек печати.'''
    target_ws = target_wb.create_sheet(title=sheet_name)

    # Копируем все ячейки
    for row in source_ws:
        for cell in row:
            target_cell = target_ws[cell.coordinate]
            target_cell.value = cell.value

            # Копируем форматирование
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

    # Копируем размеры строк и столбцов
    for row_num, row_dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dimension.height
        target_ws.row_dimensions[row_num].hidden = row_dimension.hidden

    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dimension.width
        target_ws.column_dimensions[col_letter].hidden = col_dimension.hidden

    # Копируем объединенные ячейки
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Копируем настройки печати
    copy_print_settings(source_ws, target_ws)

    return target_ws


def main():
    '''Основная функция выполнения программы.'''
    setup_logging()

    data_file = 'я. Бетон (Июль).xlsx'  # Путь к файлу с данными
    template_file = 'Шаблон.xlsx'  # Путь к шаблону для актов
    output_folder = 'Акты_бетон'  # Название для папки под акты
    output_file = 'Все_акты_бетон.xlsx'  # Итоговый файл с всеми актами

    try:
        # Проверка существования файлов
        validate_files(data_file, template_file)

        # Создаем папку для выходных файлов
        os.makedirs(output_folder, exist_ok=True)
        logging.info(f"Создана/проверена папка: {output_folder}")

        # Загрузка файлов
        wb_data = load_workbook(data_file)
        ws_data = wb_data['Бетон для АОСР']
        logging.info(f"Загружен файл данных: {data_file}")

        # Загружаем шаблон
        wb_template = load_workbook(template_file)
        ws_template = wb_template['АОСР бетон']
        logging.info(f"Загружен файл шаблона: {template_file}")

        # Создаем новую книгу для всех актов
        output_wb = Workbook()
        # Удаляем стандартный лист
        if output_wb.active:
            output_wb.remove(output_wb.active)

        output_path = os.path.join(output_folder, output_file)

        process_acts(ws_data, ws_template, output_wb, output_path)

        # Закрываем исходные файлы
        wb_data.close()
        wb_template.close()

    except FileNotFoundError as e:
        logging.error(f"Ошибка: {e}")
        print(f"Ошибка: {e}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Неожиданная ошибка: {e}")
        print(f"Неожиданная ошибка: {e}")
        sys.exit(1)


def process_acts(ws_data, ws_template, output_wb, output_path):
    '''Обработка актов из данных Excel с созданием листов в одном файле.'''
    dkbs = ('документ о качестве бетонной смеси заданного '
            'состава качества партии')
    name_uzk = ('Протокол оценки прочности бетона монолитных '
                'железобетонных конструкций')
    name_k1 = ('Акт отбора проб бетонной смеси и изготовления '
               'контрольных образцов')
    name_k2 = ('Протокол оценки прочности бетона монолитных '
               'конструкций')

    processed_count = 0

    for row in ws_data.iter_rows(min_row=3, values_only=True):
        # Пропускаем пустые строки
        if not row[0]:
            continue

        try:
            '''Сохраняем данные из файла Excel в переменные.'''
            id = row[0]
            act_number = str(row[1])
            work_name = str(row[2])
            start_date = format_date(row[3])
            end_date = format_date(row[4])
            concrete_type = str(row[5])
            mixture_number = str(row[7])
            mixture_date = format_date(row[9])
            lab_uzk = str(row[10]) if row[10] else ''
            lab_k = str(row[11]) if row[11] else ''
            lab_date = format_date(row[12])
            code = str(row[13])
            agreement_date = format_date(row[14])

            logging.info(f"Обработка акта №{id}")

            # Создаем новый лист для текущего акта
            sheet_name = f'Акт №{id}'
            ws_new = copy_worksheet(ws_template, output_wb, sheet_name)

            # Упрощенная проверка даты акта
            act_date = get_act_date(row[4], row[12], row[14])

            # Проверка ЖАН
            agreement = (f'Запись из ЖАН от {agreement_date}'
                         if agreement_date else '')

            # Проверка материалов, реестр или нет
            if mixture_number == 'Реестр':
                material1 = (f'Материалы согласно реестру '
                             f'№{act_number} от {act_date}')
                material2 = f'Реестр №{act_number} от {act_date}'
                material1_1, material2_1 = '', ''
            elif '\n' in mixture_number:
                mixture_number_parts = mixture_number.split('\n')
                mixture_date_parts = mixture_date.split('\n')
                material1 = (f'{concrete_type} - {dkbs} '
                             f'№{mixture_number_parts[0]} '
                             f'от {mixture_date_parts[0]}')
                material1_1 = (f'{concrete_type} - {dkbs} '
                               f'№{mixture_number_parts[1]} '
                               f'от {mixture_date_parts[1]}')
                material2 = (f'{dkbs.capitalize()} '
                             f'№{mixture_number_parts[0]} '
                             f'от {mixture_date_parts[0]}')
                material2_1 = (f'{dkbs.capitalize()} '
                               f'№{mixture_number_parts[1]} '
                               f'от {mixture_date_parts[1]}')
            else:
                material1 = (f'{concrete_type} - {dkbs} '
                             f'№{mixture_number} от {mixture_date}')
                material2 = (f'{dkbs.capitalize()} '
                             f'№{mixture_number} от {mixture_date}')
                material1_1, material2_1 = '', ''

            # Проверка лаборатории, УЗК или К
            if lab_uzk:
                lab1 = (f'{name_uzk} №{lab_uzk}-УЗК/2/1.3В-2025 '
                        f'от {lab_date}')
                lab2 = ''
            else:
                lab1 = (f'{name_k1} №{lab_k}-К/2/1.3В-2025 '
                        f'от {start_date}')
                lab2 = (f'{name_k2} №{lab_k}-К7/2/1.3В-2025 '
                        f'от {lab_date}')

            # Список соответствия мест в Excel и переменных
            replacements = {
                '[№ акта]': act_number,
                '[Наименование работ]': work_name,
                '[Дата начала работы]': start_date,
                '[Дата окончания работы]': end_date,
                '[Шифр]': code,
                '[Согласование]': agreement,
                '[Материалы1]': material1,
                '[Материалы2]': material2,
                '[Материалы1_1]': material1_1,
                '[Материалы2_1]': material2_1,
                '[Лаборатория1]': lab1,
                '[Лаборатория2]': lab2,
                '[Дата акта]': act_date
            }

            # Заполнение ячеек в новом листе
            for row_cells in ws_new:
                for cell in row_cells:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        for placeholder, value in replacements.items():
                            if placeholder in cell_value:
                                cell_value = cell_value.replace(
                                    placeholder, value)
                        cell.value = cell_value

            # Скрытие строк по известным номерам
            rows_to_hide = []

            if not material1_1:
                rows_to_hide.append(76)
            if not material2_1:
                rows_to_hide.append(97)
            if not lab2:
                rows_to_hide.append(81)
                rows_to_hide.append(99)
            if not agreement_date:
                rows_to_hide.append(100)

            for row_num in rows_to_hide:
                ws_new.row_dimensions[row_num].hidden = True

            processed_count += 1
            logging.info(f"Успешно создан лист '{sheet_name}'")

        except Exception as e:
            row_id = row[0] if row[0] else 'Неизвестно'
            msg = f"Ошибка при обработке акта №{row_id}: {e}"
            logging.error(msg)
            continue

    # Сохраняем финальный файл со всеми актами
    output_wb.save(output_path)
    logging.info(f"Обработка завершена. Создано актов: {processed_count}")
    logging.info(f"Файл сохранен: {output_path}")
    print(f"Обработка завершена! Создано актов: {processed_count}")
    print(f"Все акты сохранены в файл: {output_path}")


if __name__ == "__main__":
    main()
